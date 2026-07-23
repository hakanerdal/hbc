#!/usr/bin/env python3
"""hükümler_tablosu.xlsx → mock-hukumler-tablosu.js"""
import json
import sys
from pathlib import Path

import openpyxl

DEFAULT_XLSX = Path.home() / "OneDrive" / "Masaüstü" / "yazı oluştur" / "hükümler_tablosu.xlsx"
OUT_JS = (
    Path(__file__).resolve().parents[1]
    / "calismalarim"
    / "yazi_olustur"
    / "mockup"
    / "mock-hukumler-tablosu.js"
)


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        raise SystemExit(f"Excel bulunamadı: {xlsx}")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            ka = ws.cell(r, 1).value
            sk = ws.cell(r, 2).value
            md = ws.cell(r, 3).value
            if not ka or not sk or not md:
                continue
            rows.append(
                {
                    "id": len(rows) + 1,
                    "kaynak": str(sheet).strip(),
                    "koruma_alani": str(ka).strip(),
                    "sektor": str(sk).strip(),
                    "madde": str(md).strip(),
                }
            )

    OUT_JS.parent.mkdir(parents=True, exist_ok=True)
    with OUT_JS.open("w", encoding="utf-8") as f:
        f.write("/** hükümler_tablosu.xlsx — otomatik üretildi */\n")
        f.write("window.HUKUMLER_TABLOSU = ")
        json.dump(rows, f, ensure_ascii=False)
        f.write(";\n")
    print("Yazıldı:", OUT_JS, f"({len(rows)} kayıt)")


if __name__ == "__main__":
    main()
