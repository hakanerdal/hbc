#!/usr/bin/env python3
"""Excel Yazı hazırlama programı — Seçenekler sayfasını mock-yazi-excel-secenekler.js üretir."""
import json
import sys
from pathlib import Path

import openpyxl

DEFAULT_XLSM = Path.home() / "OneDrive" / "Masaüstü" / "yazı oluştur" / "Yazı hazırlama programı_v.46.xlsm"
OUT_JS = Path(__file__).resolve().parents[1] / "calismalarim" / "yazi_olustur" / "mockup" / "mock-yazi-excel-secenekler.js"


def col_vals(ws, col: int, start_row: int = 4) -> list:
    """Seçenekler sayfasında satır 3 = alan etiketi (seçenek değil); seçenekler satır 4'ten başlar."""
    out = []
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None or str(v).strip() == "":
            continue
        if isinstance(v, float) and v == int(v):
            v = int(v)
        s = str(v).strip()
        if s.startswith("!!!"):
            continue
        out.append(v)
    return out


def main() -> None:
    xlsm = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSM
    if not xlsm.is_file():
        raise SystemExit(f"Excel bulunamadı: {xlsm}")

    wb = openpyxl.load_workbook(xlsm, data_only=True)
    ws_opt = next(wb[s] for s in wb.sheetnames if "enek" in s)
    ws_faal = wb["Faaliyet"]

    field_map = {
        1: "havzada_mi",
        2: "ek",
        3: "baraj_gol",
        4: "koruma_plani_yili",
        5: "suki",
        6: "tahsis_satis",
        7: "il",
        8: "hitap",
        9: "koruma_planlari",
        10: "taskin_gorusu",
        11: "koruma_alani_mesafe",
    }
    data = {field_map[c]: col_vals(ws_opt, c) for c in field_map}
    data["talep_turu"] = [
        ws_faal.cell(r, 1).value for r in range(2, ws_faal.max_row + 1) if ws_faal.cell(r, 1).value
    ]
    data["faaliyet_tedbir"] = {}
    for r in range(2, ws_faal.max_row + 1):
        talep = ws_faal.cell(r, 1).value
        tedbir = ws_faal.cell(r, 2).value
        if talep is None or str(talep).strip() == "":
            continue
        key = str(talep).strip()
        data["faaliyet_tedbir"][key] = str(tedbir).strip() if tedbir is not None else ""
    data["evet_hayir"] = ["Evet", "Hayır"]
    data["ilgide_suki"] = ["Evet", "Hayır"]
    data["hukum_yaz"] = ["Evet", "Hayır"]
    data["tarim_mudurlugu"] = ["Yok", "Var"]
    data["ilgide_kurum"] = ["Evet", "Hayır"]
    data["suki_yonetmelik"] = ["Evet", "Hayır"]
    data["yeralti_suyu"] = ["Evet", "Hayır"]

    OUT_JS.parent.mkdir(parents=True, exist_ok=True)
    with OUT_JS.open("w", encoding="utf-8") as f:
        f.write("/** Excel Yazı hazırlama programı — Seçenekler (otomatik üretildi) */\n")
        f.write("window.YAZI_EXCEL_SECENEKLER = ")
        json.dump(data, f, ensure_ascii=False)
        f.write(";\n")
    print("Yazıldı:", OUT_JS)


if __name__ == "__main__":
    main()
